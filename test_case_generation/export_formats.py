"""
Export Formats for Healthcare Test Cases

Supports export to various enterprise tools including Jira, Azure DevOps,
Excel spreadsheets, and JSON formats for integration with test management systems.
"""

import logging
from typing import Dict, List, Optional, Any, Union
import json
import csv
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import requests
except ImportError:
    requests = None

logger = logging.getLogger(__name__)


class ExportManager:
    """Manages export of test cases to various formats."""
    
    def __init__(self):
        """Initialize export manager."""
        self.supported_formats = ['json', 'excel', 'csv', 'jira', 'azure_devops']
        
    def export_test_cases(self, test_cases: List[Any], 
                         output_path: Union[str, Path], 
                         format_type: str = 'json',
                         **kwargs) -> bool:
        """
        Export test cases to specified format.
        
        Args:
            test_cases: List of test cases to export
            output_path: Path for output file
            format_type: Export format (json, excel, csv, jira, azure_devops)
            **kwargs: Additional format-specific parameters
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            if format_type == 'json':
                return self._export_to_json(test_cases, output_path, **kwargs)
            elif format_type == 'excel':
                return self._export_to_excel(test_cases, output_path, **kwargs)
            elif format_type == 'csv':
                return self._export_to_csv(test_cases, output_path, **kwargs)
            elif format_type == 'jira':
                return self._export_to_jira(test_cases, output_path, **kwargs)
            elif format_type == 'azure_devops':
                return self._export_to_azure_devops(test_cases, output_path, **kwargs)
            else:
                logger.error(f"Unsupported export format: {format_type}")
                return False
                
        except Exception as e:
            logger.error(f"Export failed: {str(e)}")
            return False
            
    def _export_to_json(self, test_cases: List[Any], output_path: Union[str, Path], 
                       **kwargs) -> bool:
        """Export test cases to JSON format."""
        output_path = Path(output_path)
        
        # Convert test cases to dictionary format
        export_data = {
            'export_info': {
                'export_timestamp': datetime.now().isoformat(),
                'total_test_cases': len(test_cases),
                'export_format': 'json',
                'version': '1.0'
            },
            'test_cases': []
        }
        
        for tc in test_cases:
            tc_dict = self._test_case_to_dict(tc)
            export_data['test_cases'].append(tc_dict)
            
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
            
        logger.info(f"Exported {len(test_cases)} test cases to JSON: {output_path}")
        return True
        
    def _export_to_excel(self, test_cases: List[Any], output_path: Union[str, Path], 
                        **kwargs) -> bool:
        """Export test cases to Excel format."""
        output_path = Path(output_path)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create test cases sheet
        self._create_test_cases_sheet(wb, test_cases)
        
        # Create summary sheet
        self._create_summary_sheet(wb, test_cases)
        
        # Create traceability sheet
        self._create_traceability_sheet(wb, test_cases)
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"Exported {len(test_cases)} test cases to Excel: {output_path}")
        return True
        
    def _create_test_cases_sheet(self, wb: Workbook, test_cases: List[Any]) -> None:
        """Create test cases sheet in Excel workbook."""
        ws = wb.create_sheet("Test Cases")
        
        # Define headers
        headers = [
            'Test Case ID', 'Title', 'Description', 'Type', 'Priority', 'Requirement ID',
            'Compliance Refs', 'Prerequisites', 'Test Steps', 'Expected Outcome',
            'Pass Criteria', 'Fail Criteria', 'Estimated Duration (min)', 'Created Date'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # Write test case data
        for row, tc in enumerate(test_cases, 2):
            ws.cell(row=row, column=1, value=getattr(tc, 'id', ''))
            ws.cell(row=row, column=2, value=getattr(tc, 'title', ''))
            ws.cell(row=row, column=3, value=getattr(tc, 'description', ''))
            ws.cell(row=row, column=4, value=getattr(tc, 'test_case_type', '').value if hasattr(getattr(tc, 'test_case_type', ''), 'value') else str(getattr(tc, 'test_case_type', '')))
            ws.cell(row=row, column=5, value=getattr(tc, 'priority', '').value if hasattr(getattr(tc, 'priority', ''), 'value') else str(getattr(tc, 'priority', '')))
            ws.cell(row=row, column=6, value=getattr(tc, 'requirement_id', ''))
            ws.cell(row=row, column=7, value=', '.join(getattr(tc, 'compliance_refs', [])))
            ws.cell(row=row, column=8, value='; '.join(getattr(tc, 'prerequisites', [])))
            
            # Format test steps
            test_steps = getattr(tc, 'test_steps', [])
            steps_text = []
            for step in test_steps:
                if hasattr(step, 'step_number') and hasattr(step, 'action'):
                    steps_text.append(f"{step.step_number}. {step.action}")
            ws.cell(row=row, column=9, value='; '.join(steps_text))
            
            ws.cell(row=row, column=10, value=getattr(tc, 'expected_outcome', ''))
            ws.cell(row=row, column=11, value='; '.join(getattr(tc, 'pass_criteria', [])))
            ws.cell(row=row, column=12, value='; '.join(getattr(tc, 'fail_criteria', [])))
            ws.cell(row=row, column=13, value=getattr(tc, 'estimated_duration', ''))
            ws.cell(row=row, column=14, value=getattr(tc, 'created_date', ''))
            
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _create_summary_sheet(self, wb: Workbook, test_cases: List[Any]) -> None:
        """Create summary sheet in Excel workbook."""
        ws = wb.create_sheet("Summary")
        
        # Calculate summary statistics
        total_tc = len(test_cases)
        by_type = {}
        by_priority = {}
        
        for tc in test_cases:
            tc_type = getattr(tc, 'test_case_type', '').value if hasattr(getattr(tc, 'test_case_type', ''), 'value') else str(getattr(tc, 'test_case_type', ''))
            priority = getattr(tc, 'priority', '').value if hasattr(getattr(tc, 'priority', ''), 'value') else str(getattr(tc, 'priority', ''))
            
            by_type[tc_type] = by_type.get(tc_type, 0) + 1
            by_priority[priority] = by_priority.get(priority, 0) + 1
            
        # Write summary data
        ws.cell(row=1, column=1, value="Test Case Summary").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value="Total Test Cases:").font = Font(bold=True)
        ws.cell(row=2, column=2, value=total_tc)
        
        ws.cell(row=4, column=1, value="By Type:").font = Font(bold=True)
        row = 5
        for tc_type, count in by_type.items():
            ws.cell(row=row, column=1, value=tc_type)
            ws.cell(row=row, column=2, value=count)
            row += 1
            
        ws.cell(row=row+1, column=1, value="By Priority:").font = Font(bold=True)
        row += 2
        for priority, count in by_priority.items():
            ws.cell(row=row, column=1, value=priority)
            ws.cell(row=row, column=2, value=count)
            row += 1
            
    def _create_traceability_sheet(self, wb: Workbook, test_cases: List[Any]) -> None:
        """Create traceability matrix sheet in Excel workbook."""
        ws = wb.create_sheet("Traceability Matrix")
        
        # Headers
        headers = ['Test Case ID', 'Requirement ID', 'Compliance References', 'Test Type', 'Priority']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
        # Data
        for row, tc in enumerate(test_cases, 2):
            ws.cell(row=row, column=1, value=getattr(tc, 'id', ''))
            ws.cell(row=row, column=2, value=getattr(tc, 'requirement_id', ''))
            ws.cell(row=row, column=3, value=', '.join(getattr(tc, 'compliance_refs', [])))
            ws.cell(row=row, column=4, value=getattr(tc, 'test_case_type', '').value if hasattr(getattr(tc, 'test_case_type', ''), 'value') else str(getattr(tc, 'test_case_type', '')))
            ws.cell(row=row, column=5, value=getattr(tc, 'priority', '').value if hasattr(getattr(tc, 'priority', ''), 'value') else str(getattr(tc, 'priority', '')))
            
    def _export_to_csv(self, test_cases: List[Any], output_path: Union[str, Path], 
                      **kwargs) -> bool:
        """Export test cases to CSV format."""
        output_path = Path(output_path)
        
        # Convert test cases to list of dictionaries
        csv_data = []
        for tc in test_cases:
            tc_dict = self._test_case_to_dict(tc)
            csv_data.append(tc_dict)
            
        # Write to CSV
        if csv_data:
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=csv_data[0].keys())
                writer.writeheader()
                writer.writerows(csv_data)
                
        logger.info(f"Exported {len(test_cases)} test cases to CSV: {output_path}")
        return True
        
    def _export_to_jira(self, test_cases: List[Any], output_path: Union[str, Path], 
                       **kwargs) -> bool:
        """Export test cases to Jira format."""
        output_path = Path(output_path)
        
        # Jira-specific parameters
        project_key = kwargs.get('project_key', 'TEST')
        issue_type = kwargs.get('issue_type', 'Test')
        jira_url = kwargs.get('jira_url', '')
        username = kwargs.get('username', '')
        api_token = kwargs.get('api_token', '')
        
        if not jira_url or not username or not api_token:
            logger.warning("Jira credentials not provided. Creating JSON file for manual import.")
            return self._export_jira_json(test_cases, output_path, project_key, issue_type)
            
        # Export to Jira via API
        return self._export_to_jira_api(test_cases, jira_url, username, api_token, 
                                      project_key, issue_type)
        
    def _export_jira_json(self, test_cases: List[Any], output_path: Path, 
                         project_key: str, issue_type: str) -> bool:
        """Export test cases as Jira-compatible JSON."""
        jira_data = {
            'issues': []
        }
        
        for tc in test_cases:
            # Create Jira issue
            issue = {
                'fields': {
                    'project': {'key': project_key},
                    'issuetype': {'name': issue_type},
                    'summary': getattr(tc, 'title', ''),
                    'description': self._format_jira_description(tc),
                    'priority': {'name': self._map_priority_to_jira(getattr(tc, 'priority', ''))},
                    'labels': ['healthcare', 'test-case', 'ai-generated'],
                    'customfield_10001': getattr(tc, 'requirement_id', ''),  # Requirement ID
                    'customfield_10002': ', '.join(getattr(tc, 'compliance_refs', [])),  # Compliance Refs
                }
            }
            
            # Add test steps as subtasks or comments
            test_steps = getattr(tc, 'test_steps', [])
            if test_steps:
                steps_text = "Test Steps:\\n"
                for step in test_steps:
                    if hasattr(step, 'step_number') and hasattr(step, 'action'):
                        steps_text += f"* {step.step_number}. {step.action}\\n"
                issue['fields']['description'] += f"\\n\\n{steps_text}"
                
            jira_data['issues'].append(issue)
            
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(jira_data, f, indent=2, ensure_ascii=False)
            
        logger.info(f"Exported {len(test_cases)} test cases to Jira JSON: {output_path}")
        return True
        
    def _export_to_jira_api(self, test_cases: List[Any], jira_url: str, username: str, 
                           api_token: str, project_key: str, issue_type: str) -> bool:
        """Export test cases to Jira via API."""
        if not requests:
            logger.error("Requests library not available for Jira API export")
            return False
            
        # Jira API endpoint
        api_url = f"{jira_url}/rest/api/2/issue"
        
        # Headers for authentication
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Basic {self._encode_credentials(username, api_token)}'
        }
        
        success_count = 0
        
        for tc in test_cases:
            try:
                # Create issue data
                issue_data = {
                    'fields': {
                        'project': {'key': project_key},
                        'issuetype': {'name': issue_type},
                        'summary': getattr(tc, 'title', ''),
                        'description': self._format_jira_description(tc),
                        'priority': {'name': self._map_priority_to_jira(getattr(tc, 'priority', ''))},
                        'labels': ['healthcare', 'test-case', 'ai-generated']
                    }
                }
                
                # Send request
                response = requests.post(api_url, json=issue_data, headers=headers)
                
                if response.status_code == 201:
                    success_count += 1
                    logger.info(f"Created Jira issue for test case: {getattr(tc, 'id', '')}")
                else:
                    logger.error(f"Failed to create Jira issue: {response.text}")
                    
            except Exception as e:
                logger.error(f"Error creating Jira issue: {str(e)}")
                
        logger.info(f"Successfully exported {success_count}/{len(test_cases)} test cases to Jira")
        return success_count > 0
        
    def _export_to_azure_devops(self, test_cases: List[Any], output_path: Union[str, Path], 
                               **kwargs) -> bool:
        """Export test cases to Azure DevOps format."""
        output_path = Path(output_path)
        
        # Azure DevOps parameters
        organization = kwargs.get('organization', '')
        project = kwargs.get('project', '')
        personal_access_token = kwargs.get('personal_access_token', '')
        
        if not organization or not project or not personal_access_token:
            logger.warning("Azure DevOps credentials not provided. Creating JSON file for manual import.")
            return self._export_azure_devops_json(test_cases, output_path, project)
            
        # Export to Azure DevOps via API
        return self._export_to_azure_devops_api(test_cases, organization, project, 
                                              personal_access_token)
        
    def _export_azure_devops_json(self, test_cases: List[Any], output_path: Path, 
                                 project: str) -> bool:
        """Export test cases as Azure DevOps-compatible JSON."""
        azure_data = {
            'testCases': []
        }
        
        for tc in test_cases:
            # Create Azure DevOps test case
            test_case = {
                'name': getattr(tc, 'title', ''),
                'description': getattr(tc, 'description', ''),
                'priority': self._map_priority_to_azure_devops(getattr(tc, 'priority', '')),
                'tags': ['healthcare', 'test-case', 'ai-generated'],
                'fields': {
                    'System.AreaPath': project,
                    'System.IterationPath': project,
                    'Custom.RequirementId': getattr(tc, 'requirement_id', ''),
                    'Custom.ComplianceRefs': ', '.join(getattr(tc, 'compliance_refs', []))
                },
                'steps': []
            }
            
            # Add test steps
            test_steps = getattr(tc, 'test_steps', [])
            for step in test_steps:
                if hasattr(step, 'step_number') and hasattr(step, 'action'):
                    step_data = {
                        'id': step.step_number,
                        'action': step.action,
                        'expectedResult': getattr(step, 'expected_result', '')
                    }
                    test_case['steps'].append(step_data)
                    
            azure_data['testCases'].append(test_case)
            
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(azure_data, f, indent=2, ensure_ascii=False)
            
        logger.info(f"Exported {len(test_cases)} test cases to Azure DevOps JSON: {output_path}")
        return True
        
    def _export_to_azure_devops_api(self, test_cases: List[Any], organization: str, 
                                   project: str, personal_access_token: str) -> bool:
        """Export test cases to Azure DevOps via API."""
        if not requests:
            logger.error("Requests library not available for Azure DevOps API export")
            return False
            
        # Azure DevOps API endpoint
        api_url = f"https://dev.azure.com/{organization}/{project}/_apis/test/plans"
        
        # Headers for authentication
        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Basic {self._encode_credentials("", personal_access_token)}'
        }
        
        # This is a simplified implementation
        # In practice, you would need to create test plans, test suites, and test cases
        logger.info("Azure DevOps API export requires complex setup with test plans and suites")
        return False
        
    def _test_case_to_dict(self, test_case: Any) -> Dict[str, Any]:
        """Convert test case object to dictionary."""
        tc_dict = {
            'id': getattr(test_case, 'id', ''),
            'title': getattr(test_case, 'title', ''),
            'description': getattr(test_case, 'description', ''),
            'test_case_type': getattr(test_case, 'test_case_type', '').value if hasattr(getattr(test_case, 'test_case_type', ''), 'value') else str(getattr(test_case, 'test_case_type', '')),
            'priority': getattr(test_case, 'priority', '').value if hasattr(getattr(test_case, 'priority', ''), 'value') else str(getattr(test_case, 'priority', '')),
            'requirement_id': getattr(test_case, 'requirement_id', ''),
            'compliance_refs': getattr(test_case, 'compliance_refs', []),
            'prerequisites': getattr(test_case, 'prerequisites', []),
            'expected_outcome': getattr(test_case, 'expected_outcome', ''),
            'pass_criteria': getattr(test_case, 'pass_criteria', []),
            'fail_criteria': getattr(test_case, 'fail_criteria', []),
            'estimated_duration': getattr(test_case, 'estimated_duration', ''),
            'created_date': getattr(test_case, 'created_date', ''),
            'last_modified': getattr(test_case, 'last_modified', ''),
            'test_steps': []
        }
        
        # Add test steps
        test_steps = getattr(test_case, 'test_steps', [])
        for step in test_steps:
            step_dict = {
                'step_number': getattr(step, 'step_number', ''),
                'action': getattr(step, 'action', ''),
                'expected_result': getattr(step, 'expected_result', ''),
                'data_inputs': getattr(step, 'data_inputs', {}),
                'preconditions': getattr(step, 'preconditions', ''),
                'postconditions': getattr(step, 'postconditions', '')
            }
            tc_dict['test_steps'].append(step_dict)
            
        return tc_dict
        
    def _format_jira_description(self, test_case: Any) -> str:
        """Format test case description for Jira."""
        description = f"*Description:* {getattr(test_case, 'description', '')}\\n\\n"
        
        if getattr(test_case, 'prerequisites', []):
            description += f"*Prerequisites:*\\n"
            for prereq in getattr(test_case, 'prerequisites', []):
                description += f"* {prereq}\\n"
            description += "\\n"
            
        if getattr(test_case, 'expected_outcome', ''):
            description += f"*Expected Outcome:* {getattr(test_case, 'expected_outcome', '')}\\n\\n"
            
        if getattr(test_case, 'pass_criteria', []):
            description += f"*Pass Criteria:*\\n"
            for criteria in getattr(test_case, 'pass_criteria', []):
                description += f"* {criteria}\\n"
            description += "\\n"
            
        if getattr(test_case, 'fail_criteria', []):
            description += f"*Fail Criteria:*\\n"
            for criteria in getattr(test_case, 'fail_criteria', []):
                description += f"* {criteria}\\n"
                
        return description
        
    def _map_priority_to_jira(self, priority: Any) -> str:
        """Map test case priority to Jira priority."""
        priority_mapping = {
            'critical': 'Highest',
            'high': 'High',
            'medium': 'Medium',
            'low': 'Low'
        }
        
        priority_str = priority.value if hasattr(priority, 'value') else str(priority)
        return priority_mapping.get(priority_str.lower(), 'Medium')
        
    def _map_priority_to_azure_devops(self, priority: Any) -> int:
        """Map test case priority to Azure DevOps priority (1-4 scale)."""
        priority_mapping = {
            'critical': 1,
            'high': 2,
            'medium': 3,
            'low': 4
        }
        
        priority_str = priority.value if hasattr(priority, 'value') else str(priority)
        return priority_mapping.get(priority_str.lower(), 3)
        
    def _encode_credentials(self, username: str, password: str) -> str:
        """Encode credentials for basic authentication."""
        import base64
        credentials = f"{username}:{password}"
        encoded_credentials = base64.b64encode(credentials.encode()).decode()
        return encoded_credentials

