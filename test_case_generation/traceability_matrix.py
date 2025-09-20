"""
Traceability Matrix Generator for Healthcare Test Cases

Generates comprehensive traceability matrices linking requirements to test cases
and compliance standards for healthcare software validation.
"""

import logging
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


@dataclass
class TraceabilityItem:
    """Individual traceability item."""
    requirement_id: str
    requirement_description: str
    test_case_id: str
    test_case_title: str
    compliance_standard: str
    traceability_level: str  # direct, indirect, related
    coverage_status: str  # covered, partial, not_covered
    evidence_required: List[str]


class TraceabilityMatrixGenerator:
    """Generates traceability matrices for healthcare test cases."""
    
    def __init__(self):
        """Initialize traceability matrix generator."""
        self.compliance_standards = [
            'FDA_21_CFR_820', 'FDA_21_CFR_11', 'ISO_13485', 'IEC_62304', 
            'GDPR', 'HIPAA', 'IEC_60601', 'IEC_62366'
        ]
        
    def generate_traceability_matrix(self, requirements: List[Any], 
                                   test_cases: List[Any], 
                                   compliance_mappings: List[Any] = None) -> Dict[str, Any]:
        """
        Generate comprehensive traceability matrix.
        
        Args:
            requirements: List of requirements
            test_cases: List of test cases
            compliance_mappings: Optional compliance mappings
            
        Returns:
            Traceability matrix dictionary
        """
        # Create traceability items
        traceability_items = self._create_traceability_items(
            requirements, test_cases, compliance_mappings
        )
        
        # Generate different views of the matrix
        matrix_views = {
            'requirement_to_test_case': self._create_requirement_to_test_case_view(traceability_items),
            'test_case_to_requirement': self._create_test_case_to_requirement_view(traceability_items),
            'compliance_coverage': self._create_compliance_coverage_view(traceability_items),
            'coverage_summary': self._create_coverage_summary(traceability_items, requirements, test_cases)
        }
        
        return {
            'traceability_items': [item.__dict__ for item in traceability_items],
            'matrix_views': matrix_views,
            'generation_timestamp': self._get_current_timestamp()
        }
        
    def _create_traceability_items(self, requirements: List[Any], 
                                 test_cases: List[Any], 
                                 compliance_mappings: List[Any] = None) -> List[TraceabilityItem]:
        """Create traceability items from requirements and test cases."""
        traceability_items = []
        
        # Create mapping from requirement ID to compliance standards
        req_compliance_map = {}
        if compliance_mappings:
            for mapping in compliance_mappings:
                req_id = getattr(mapping, 'requirement_id', '')
                standard = getattr(mapping, 'standard', '')
                if req_id not in req_compliance_map:
                    req_compliance_map[req_id] = []
                req_compliance_map[req_id].append(standard)
                
        # Create traceability items for each requirement-test case pair
        for req in requirements:
            req_id = getattr(req, 'id', '')
            req_desc = getattr(req, 'description', '')
            
            # Find test cases for this requirement
            related_test_cases = [
                tc for tc in test_cases 
                if getattr(tc, 'requirement_id', '') == req_id
            ]
            
            if not related_test_cases:
                # Create item for uncovered requirement
                item = TraceabilityItem(
                    requirement_id=req_id,
                    requirement_description=req_desc,
                    test_case_id='',
                    test_case_title='',
                    compliance_standard='',
                    traceability_level='not_covered',
                    coverage_status='not_covered',
                    evidence_required=[]
                )
                traceability_items.append(item)
            else:
                # Create items for each test case
                for tc in related_test_cases:
                    tc_id = getattr(tc, 'id', '')
                    tc_title = getattr(tc, 'title', '')
                    compliance_refs = getattr(tc, 'compliance_refs', [])
                    
                    # Determine compliance standards
                    standards = req_compliance_map.get(req_id, compliance_refs)
                    if not standards:
                        standards = ['Not Specified']
                        
                    for standard in standards:
                        item = TraceabilityItem(
                            requirement_id=req_id,
                            requirement_description=req_desc,
                            test_case_id=tc_id,
                            test_case_title=tc_title,
                            compliance_standard=standard.value if hasattr(standard, 'value') else str(standard),
                            traceability_level=self._determine_traceability_level(req_desc, tc_title),
                            coverage_status=self._determine_coverage_status(req_desc, tc_title),
                            evidence_required=self._get_evidence_requirements(standard)
                        )
                        traceability_items.append(item)
                        
        return traceability_items
        
    def _create_requirement_to_test_case_view(self, traceability_items: List[TraceabilityItem]) -> Dict[str, Any]:
        """Create requirement-to-test-case view of traceability matrix."""
        view = {}
        
        for item in traceability_items:
            req_id = item.requirement_id
            if req_id not in view:
                view[req_id] = {
                    'requirement_description': item.requirement_description,
                    'test_cases': [],
                    'compliance_standards': set(),
                    'coverage_status': 'not_covered'
                }
                
            if item.test_case_id:
                view[req_id]['test_cases'].append({
                    'test_case_id': item.test_case_id,
                    'test_case_title': item.test_case_title,
                    'compliance_standard': item.compliance_standard,
                    'traceability_level': item.traceability_level,
                    'coverage_status': item.coverage_status
                })
                
            if item.compliance_standard:
                view[req_id]['compliance_standards'].add(item.compliance_standard)
                
            # Update coverage status
            if item.coverage_status == 'covered':
                view[req_id]['coverage_status'] = 'covered'
            elif item.coverage_status == 'partial' and view[req_id]['coverage_status'] != 'covered':
                view[req_id]['coverage_status'] = 'partial'
                
        # Convert sets to lists for JSON serialization
        for req_data in view.values():
            req_data['compliance_standards'] = list(req_data['compliance_standards'])
            
        return view
        
    def _create_test_case_to_requirement_view(self, traceability_items: List[TraceabilityItem]) -> Dict[str, Any]:
        """Create test-case-to-requirement view of traceability matrix."""
        view = {}
        
        for item in traceability_items:
            if not item.test_case_id:
                continue
                
            tc_id = item.test_case_id
            if tc_id not in view:
                view[tc_id] = {
                    'test_case_title': item.test_case_title,
                    'requirements': [],
                    'compliance_standards': set()
                }
                
            view[tc_id]['requirements'].append({
                'requirement_id': item.requirement_id,
                'requirement_description': item.requirement_description,
                'compliance_standard': item.compliance_standard,
                'traceability_level': item.traceability_level,
                'coverage_status': item.coverage_status
            })
            
            if item.compliance_standard:
                view[tc_id]['compliance_standards'].add(item.compliance_standard)
                
        # Convert sets to lists for JSON serialization
        for tc_data in view.values():
            tc_data['compliance_standards'] = list(tc_data['compliance_standards'])
            
        return view
        
    def _create_compliance_coverage_view(self, traceability_items: List[TraceabilityItem]) -> Dict[str, Any]:
        """Create compliance coverage view of traceability matrix."""
        view = {}
        
        for item in traceability_items:
            standard = item.compliance_standard
            if not standard or standard == 'Not Specified':
                continue
                
            if standard not in view:
                view[standard] = {
                    'requirements': set(),
                    'test_cases': set(),
                    'coverage_percentage': 0.0
                }
                
            if item.requirement_id:
                view[standard]['requirements'].add(item.requirement_id)
            if item.test_case_id:
                view[standard]['test_cases'].add(item.test_case_id)
                
        # Calculate coverage percentages
        for standard, data in view.items():
            req_count = len(data['requirements'])
            tc_count = len(data['test_cases'])
            
            if req_count > 0:
                data['coverage_percentage'] = (tc_count / req_count) * 100
            else:
                data['coverage_percentage'] = 0.0
                
            # Convert sets to lists for JSON serialization
            data['requirements'] = list(data['requirements'])
            data['test_cases'] = list(data['test_cases'])
            
        return view
        
    def _create_coverage_summary(self, traceability_items: List[TraceabilityItem], 
                               requirements: List[Any], test_cases: List[Any]) -> Dict[str, Any]:
        """Create coverage summary statistics."""
        total_requirements = len(requirements)
        total_test_cases = len(test_cases)
        
        # Count covered requirements
        covered_requirements = set()
        for item in traceability_items:
            if item.test_case_id and item.coverage_status in ['covered', 'partial']:
                covered_requirements.add(item.requirement_id)
                
        covered_count = len(covered_requirements)
        coverage_percentage = (covered_count / total_requirements * 100) if total_requirements > 0 else 0.0
        
        # Count by traceability level
        traceability_levels = {}
        for item in traceability_items:
            level = item.traceability_level
            traceability_levels[level] = traceability_levels.get(level, 0) + 1
            
        # Count by compliance standard
        compliance_coverage = {}
        for item in traceability_items:
            standard = item.compliance_standard
            if standard and standard != 'Not Specified':
                if standard not in compliance_coverage:
                    compliance_coverage[standard] = {'requirements': 0, 'test_cases': 0}
                if item.requirement_id:
                    compliance_coverage[standard]['requirements'] += 1
                if item.test_case_id:
                    compliance_coverage[standard]['test_cases'] += 1
                    
        return {
            'total_requirements': total_requirements,
            'total_test_cases': total_test_cases,
            'covered_requirements': covered_count,
            'coverage_percentage': round(coverage_percentage, 2),
            'traceability_levels': traceability_levels,
            'compliance_coverage': compliance_coverage
        }
        
    def _determine_traceability_level(self, requirement_desc: str, test_case_title: str) -> str:
        """Determine traceability level between requirement and test case."""
        req_lower = requirement_desc.lower()
        tc_lower = test_case_title.lower()
        
        # Direct traceability indicators
        direct_indicators = [
            'verify', 'test', 'validate', 'check', 'ensure'
        ]
        
        if any(indicator in tc_lower for indicator in direct_indicators):
            return 'direct'
            
        # Check for keyword overlap
        req_words = set(req_lower.split())
        tc_words = set(tc_lower.split())
        overlap = len(req_words.intersection(tc_words))
        
        if overlap >= 3:
            return 'direct'
        elif overlap >= 1:
            return 'indirect'
        else:
            return 'related'
            
    def _determine_coverage_status(self, requirement_desc: str, test_case_title: str) -> str:
        """Determine coverage status of requirement by test case."""
        req_lower = requirement_desc.lower()
        tc_lower = test_case_title.lower()
        
        # Check for comprehensive coverage
        coverage_indicators = [
            'positive', 'negative', 'boundary', 'integration', 'performance'
        ]
        
        if any(indicator in tc_lower for indicator in coverage_indicators):
            return 'covered'
            
        # Check for partial coverage
        partial_indicators = [
            'basic', 'simple', 'minimal', 'initial'
        ]
        
        if any(indicator in tc_lower for indicator in partial_indicators):
            return 'partial'
            
        return 'covered'  # Default to covered if test case exists
        
    def _get_evidence_requirements(self, compliance_standard: Any) -> List[str]:
        """Get evidence requirements for compliance standard."""
        evidence_map = {
            'FDA_21_CFR_820': [
                'Design and development documentation',
                'Risk management file',
                'Verification and validation records'
            ],
            'FDA_21_CFR_11': [
                'System validation documentation',
                'Electronic signature procedures',
                'Audit trail records'
            ],
            'ISO_13485': [
                'Quality management system documentation',
                'Risk management documentation',
                'Design and development records'
            ],
            'IEC_62304': [
                'Software development plan',
                'Software requirements specification',
                'Software verification and validation records'
            ],
            'GDPR': [
                'Data protection impact assessment',
                'Privacy by design documentation',
                'Consent management records'
            ],
            'HIPAA': [
                'Risk assessment documentation',
                'Security policies and procedures',
                'Access control documentation'
            ],
            'IEC_60601': [
                'Risk management file',
                'Essential performance documentation',
                'Safety testing records'
            ],
            'IEC_62366': [
                'Usability engineering plan',
                'User interface specification',
                'Usability testing records'
            ]
        }
        
        standard_str = compliance_standard.value if hasattr(compliance_standard, 'value') else str(compliance_standard)
        return evidence_map.get(standard_str, ['General compliance documentation'])
        
    def export_traceability_matrix(self, matrix_data: Dict[str, Any], 
                                 output_path: str, format_type: str = 'excel') -> bool:
        """
        Export traceability matrix to file.
        
        Args:
            matrix_data: Traceability matrix data
            output_path: Path for output file
            format_type: Export format (excel, csv, json)
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            if format_type == 'excel':
                return self._export_to_excel(matrix_data, output_path)
            elif format_type == 'csv':
                return self._export_to_csv(matrix_data, output_path)
            elif format_type == 'json':
                return self._export_to_json(matrix_data, output_path)
            else:
                logger.error(f"Unsupported export format: {format_type}")
                return False
                
        except Exception as e:
            logger.error(f"Export failed: {str(e)}")
            return False
            
    def _export_to_excel(self, matrix_data: Dict[str, Any], output_path: str) -> bool:
        """Export traceability matrix to Excel format."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create requirement-to-test-case sheet
        self._create_requirement_to_test_case_sheet(wb, matrix_data)
        
        # Create test-case-to-requirement sheet
        self._create_test_case_to_requirement_sheet(wb, matrix_data)
        
        # Create compliance coverage sheet
        self._create_compliance_coverage_sheet(wb, matrix_data)
        
        # Create summary sheet
        self._create_summary_sheet(wb, matrix_data)
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"Exported traceability matrix to Excel: {output_path}")
        return True
        
    def _create_requirement_to_test_case_sheet(self, wb: Workbook, matrix_data: Dict[str, Any]) -> None:
        """Create requirement-to-test-case sheet."""
        ws = wb.create_sheet("Requirement to Test Case")
        
        # Headers
        headers = [
            'Requirement ID', 'Requirement Description', 'Test Case ID', 'Test Case Title',
            'Compliance Standard', 'Traceability Level', 'Coverage Status', 'Evidence Required'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
        # Data
        traceability_items = matrix_data.get('traceability_items', [])
        for row, item in enumerate(traceability_items, 2):
            ws.cell(row=row, column=1, value=item['requirement_id'])
            ws.cell(row=row, column=2, value=item['requirement_description'])
            ws.cell(row=row, column=3, value=item['test_case_id'])
            ws.cell(row=row, column=4, value=item['test_case_title'])
            ws.cell(row=row, column=5, value=item['compliance_standard'])
            ws.cell(row=row, column=6, value=item['traceability_level'])
            ws.cell(row=row, column=7, value=item['coverage_status'])
            ws.cell(row=row, column=8, value=', '.join(item['evidence_required']))
            
    def _create_test_case_to_requirement_sheet(self, wb: Workbook, matrix_data: Dict[str, Any]) -> None:
        """Create test-case-to-requirement sheet."""
        ws = wb.create_sheet("Test Case to Requirement")
        
        # Headers
        headers = [
            'Test Case ID', 'Test Case Title', 'Requirement ID', 'Requirement Description',
            'Compliance Standard', 'Traceability Level', 'Coverage Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
        # Data
        traceability_items = matrix_data.get('traceability_items', [])
        for row, item in enumerate(traceability_items, 2):
            if item['test_case_id']:  # Only include items with test cases
                ws.cell(row=row, column=1, value=item['test_case_id'])
                ws.cell(row=row, column=2, value=item['test_case_title'])
                ws.cell(row=row, column=3, value=item['requirement_id'])
                ws.cell(row=row, column=4, value=item['requirement_description'])
                ws.cell(row=row, column=5, value=item['compliance_standard'])
                ws.cell(row=row, column=6, value=item['traceability_level'])
                ws.cell(row=row, column=7, value=item['coverage_status'])
                
    def _create_compliance_coverage_sheet(self, wb: Workbook, matrix_data: Dict[str, Any]) -> None:
        """Create compliance coverage sheet."""
        ws = wb.create_sheet("Compliance Coverage")
        
        # Headers
        headers = [
            'Compliance Standard', 'Requirements Count', 'Test Cases Count', 'Coverage Percentage'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
        # Data
        compliance_coverage = matrix_data.get('matrix_views', {}).get('compliance_coverage', {})
        for row, (standard, data) in enumerate(compliance_coverage.items(), 2):
            ws.cell(row=row, column=1, value=standard)
            ws.cell(row=row, column=2, value=len(data['requirements']))
            ws.cell(row=row, column=3, value=len(data['test_cases']))
            ws.cell(row=row, column=4, value=f"{data['coverage_percentage']:.2f}%")
            
    def _create_summary_sheet(self, wb: Workbook, matrix_data: Dict[str, Any]) -> None:
        """Create summary sheet."""
        ws = wb.create_sheet("Summary")
        
        # Summary data
        coverage_summary = matrix_data.get('matrix_views', {}).get('coverage_summary', {})
        
        # Write summary
        ws.cell(row=1, column=1, value="Traceability Matrix Summary").font = Font(bold=True, size=14)
        
        ws.cell(row=3, column=1, value="Total Requirements:").font = Font(bold=True)
        ws.cell(row=3, column=2, value=coverage_summary.get('total_requirements', 0))
        
        ws.cell(row=4, column=1, value="Total Test Cases:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=coverage_summary.get('total_test_cases', 0))
        
        ws.cell(row=5, column=1, value="Covered Requirements:").font = Font(bold=True)
        ws.cell(row=5, column=2, value=coverage_summary.get('covered_requirements', 0))
        
        ws.cell(row=6, column=1, value="Coverage Percentage:").font = Font(bold=True)
        ws.cell(row=6, column=2, value=f"{coverage_summary.get('coverage_percentage', 0):.2f}%")
        
    def _export_to_csv(self, matrix_data: Dict[str, Any], output_path: str) -> bool:
        """Export traceability matrix to CSV format."""
        traceability_items = matrix_data.get('traceability_items', [])
        
        if not traceability_items:
            logger.warning("No traceability items to export")
            return False
            
        # Convert to DataFrame
        df = pd.DataFrame(traceability_items)
        
        # Export to CSV
        df.to_csv(output_path, index=False)
        
        logger.info(f"Exported traceability matrix to CSV: {output_path}")
        return True
        
    def _export_to_json(self, matrix_data: Dict[str, Any], output_path: str) -> bool:
        """Export traceability matrix to JSON format."""
        import json
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(matrix_data, f, indent=2, ensure_ascii=False)
            
        logger.info(f"Exported traceability matrix to JSON: {output_path}")
        return True
        
    def _get_current_timestamp(self) -> str:
        """Get current timestamp for matrix generation."""
        from datetime import datetime
        return datetime.now().isoformat()

